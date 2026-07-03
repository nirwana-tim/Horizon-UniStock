import json, sys, argparse
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree
from zipfile import ZipFile

FILES = {
    '1': (Path(__file__).parent / 'Dummy - Inventory Management.xlsx', 'Inventory'),
    '2': (Path(__file__).parent / 'Dummy of Freshman Seragam 2025_2026 (Responses).xlsx', 'Students'),
}

def get_sheet_names(path):
    with ZipFile(path) as z:
        tree = ElementTree.parse(z.open('xl/workbook.xml'))
    ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    return [sheet.get('name') for sheet in tree.findall('.//s:sheet', ns)]

def safe_value(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    return v

def is_valid_row(row):
    return any(v is not None for v in row)

def detect_headers(ws):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False)):
        vals = [c.value for c in row]
        if any(v is not None and isinstance(v, (str, int, float)) for v in vals):
            cleaned = []
            for v in vals:
                if isinstance(v, str) and v.strip():
                    cleaned.append(v.strip())
                elif isinstance(v, (int, float)):
                    cleaned.append(str(v))
                else:
                    cleaned.append(f'col_{len(cleaned)+1}')
            return cleaned, i + 1
    return None, None

def process_sheet(path, sheet_name, limit=0):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    headers, header_row = detect_headers(ws)
    if not headers:
        wb.close()
        return {"total_rows": 0, "headers": [], "data": []}

    data = []
    count = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not is_valid_row(row):
            continue
        if all(v is None for v in row[:len(headers)]):
            continue
        obj = {}
        for i, key in enumerate(headers):
            val = safe_value(row[i]) if i < len(row) else None
            obj[key] = val
        data.append(obj)
        count += 1
        if limit and count >= limit:
            break

    wb.close()
    return {"total_rows": len(data), "headers": headers, "data": data}

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--list', action='store_true', help='List all sheets')
    parser.add_argument('--file', choices=['1', '2'], help='File number')
    parser.add_argument('--sheet', type=str, help='Sheet name')
    parser.add_argument('--limit', type=int, default=0, help='Max rows to return')
    args = parser.parse_args()

    if args.list:
        result = {}
        for key, (path, label) in FILES.items():
            if not path.exists():
                continue
            result[str(path.name)] = get_sheet_names(path)
        print(json.dumps(result, indent=2, ensure_ascii=False))
        return

    if not args.file or not args.sheet:
        parser.print_help()
        sys.exit(1)

    path, label = FILES[args.file]
    if not path.exists():
        print(json.dumps({"error": f"File not found: {path}"}))
        sys.exit(1)

    sheet_names = get_sheet_names(path)
    if args.sheet not in sheet_names:
        print(json.dumps({"error": f"Sheet '{args.sheet}' not found. Available: {sheet_names}"}))
        sys.exit(1)

    result = process_sheet(path, args.sheet, args.limit)
    print(json.dumps(result, indent=2, ensure_ascii=False))

if __name__ == '__main__':
    main()
